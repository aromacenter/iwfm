"""Bérszámfejtési export (CSV / XLSX).

Generic column set so it can be mapped to any Hungarian payroll program
(Nexon, Kulcs-Bér, Novitax, RLB-60 …) later:

    Dolgozó neve | Adóazonosító jel | TAJ szám | Bankszámlaszám | FEOR |
    Munkakör | Heti óraszám | Beosztott óra | Ledolgozott óra |
    Szabadság nap | Betegszabadság nap | Fizetés nélküli nap | Egyéb távollét nap

Identifiers are decrypted at export time — the export endpoint is admin/manager
gated and every call is audited.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADERS = [
    "Dolgozó neve",
    "Adóazonosító jel",
    "TAJ szám",
    "Bankszámlaszám",
    "FEOR kód",
    "Munkakör",
    "Heti óraszám",
    "Beosztott óra",
    "Ledolgozott óra",
    "Bérszámfejtés alapja",
    "Fizetendő óra",
    "Szabadság (nap)",
    "Betegszabadság (nap)",
    "Fizetés nélküli (nap)",
    "Egyéb távollét (nap)",
]


@dataclass
class PayrollRow:
    name: str
    tax_id: str | None
    taj: str | None
    bank_account: str | None
    feor_code: str | None
    job_title: str | None
    weekly_hours: int
    scheduled_hours: float
    worked_hours: float
    # 'attendance' = blokkolás szerint fizetünk, 'schedule' = beosztás szerint
    payroll_source: str
    annual_leave_days: int
    sick_days: int
    unpaid_days: int
    other_days: int

    def as_list(self) -> list:
        return [
            self.name,
            self.tax_id or "",
            self.taj or "",
            self.bank_account or "",
            self.feor_code or "",
            self.job_title or "",
            self.weekly_hours,
            round(self.scheduled_hours, 2),
            round(self.worked_hours, 2),
            "Beosztás" if self.payroll_source == "schedule" else "Blokkolás",
            round(
                self.scheduled_hours if self.payroll_source == "schedule"
                else self.worked_hours,
                2,
            ),
            self.annual_leave_days,
            self.sick_days,
            self.unpaid_days,
            self.other_days,
        ]


def overlap_days(start: date, end: date, period_start: date, period_end: date) -> int:
    """Inclusive day-count of [start,end] ∩ [period_start,period_end]."""
    lo = max(start, period_start)
    hi = min(end, period_end)
    return max((hi - lo).days + 1, 0)


def build_csv(rows: list[PayrollRow], period_start: date, period_end: date) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")  # magyar Excel pontosvesszőt vár
    writer.writerow([f"Bérszámfejtési export {period_start} – {period_end}"])
    writer.writerow(HEADERS)
    for row in rows:
        writer.writerow(row.as_list())
    return buf.getvalue().encode("utf-8-sig")  # BOM az Excel ékezetekhez


def build_xlsx(rows: list[PayrollRow], period_start: date, period_end: date) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bérexport"
    ws.append([f"Bérszámfejtési export {period_start} – {period_end}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(HEADERS)
    for cell in ws[2]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row.as_list())
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 14)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
