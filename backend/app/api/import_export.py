"""Multifunkciós import/export: Excel/CSV feltöltés oszlop-hozzárendeléssel.

Folyamat:
1. GET  /entities            — mit lehet importálni/exportálni + mezőlisták
2. POST /preview             — fájl beküldése → oszlopok + mintasorok (a
                               kliens ebből építi a hozzárendelő felületet)
3. POST /import              — fájl + {mező: oszlopindex} hozzárendelés →
                               soronkénti létrehozás, hibariporttal; a névre
                               egyező meglévő rekord kihagyva (skipped)
4. GET  /export/{entity}     — teljes lista xlsx-ben

A fájl data-URL-ként (base64) érkezik — xlsx (openpyxl) és CSV támogatott.
"""

from __future__ import annotations

import base64
import csv
import io
import uuid

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Request, Response
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from datetime import UTC, datetime as dt

from app.api.deps import record_audit, require_perm
from app.db import get_db
from app.models import Asset, AssetMovement, Partner, Product, Settlement, User

router = APIRouter()

_MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB
_MAX_ROWS = 5000
_PREVIEW_ROWS = 5

PARTNER_TYPES = ("customer", "supplier", "both")

# Mező-definíciók: (kulcs, kötelező, max_hossz | None, típus)
# típus: str | int | float | partner_type
ENTITY_FIELDS: dict[str, list[dict]] = {
    "partners": [
        {"key": "name", "required": True, "max_len": 256, "type": "str"},
        {"key": "partner_type", "required": False, "max_len": 16, "type": "partner_type"},
        {"key": "tax_number", "required": False, "max_len": 32, "type": "str"},
        {"key": "eu_tax_number", "required": False, "max_len": 32, "type": "str"},
        {"key": "reg_number", "required": False, "max_len": 64, "type": "str"},
        {"key": "contact_name", "required": False, "max_len": 256, "type": "str"},
        {"key": "contact_email", "required": False, "max_len": 320, "type": "str"},
        {"key": "contact_phone", "required": False, "max_len": 32, "type": "str"},
        {"key": "website", "required": False, "max_len": 256, "type": "str"},
        {"key": "address", "required": False, "max_len": 512, "type": "str"},
        {"key": "billing_address", "required": False, "max_len": 512, "type": "str"},
        {"key": "address_zip", "required": False, "max_len": 16, "type": "str"},
        {"key": "address_city", "required": False, "max_len": 128, "type": "str"},
        {"key": "address_street", "required": False, "max_len": 256, "type": "str"},
        {"key": "address_number", "required": False, "max_len": 32, "type": "str"},
        {"key": "billing_zip", "required": False, "max_len": 16, "type": "str"},
        {"key": "billing_city", "required": False, "max_len": 128, "type": "str"},
        {"key": "billing_street", "required": False, "max_len": 256, "type": "str"},
        {"key": "billing_number", "required": False, "max_len": 32, "type": "str"},
        {"key": "bank_account", "required": False, "max_len": 64, "type": "str"},
        {"key": "payment_terms_days", "required": False, "max_len": None, "type": "int"},
        {"key": "notes", "required": False, "max_len": None, "type": "str"},
    ],
    "products": [
        {"key": "name", "required": True, "max_len": 256, "type": "str"},
        {"key": "unit", "required": False, "max_len": 16, "type": "str"},
        {"key": "grams_per_portion", "required": False, "max_len": None, "type": "int"},
        {"key": "price_per_portion", "required": False, "max_len": None, "type": "float"},
        {"key": "vat_percent", "required": False, "max_len": None, "type": "int"},
        {"key": "notes", "required": False, "max_len": None, "type": "str"},
    ],
    "assets": [
        {"key": "barcode", "required": True, "max_len": 64, "type": "str"},
        {"key": "manufacturer", "required": False, "max_len": 128, "type": "str"},
        {"key": "name", "required": True, "max_len": 256, "type": "str"},  # Típus
        {"key": "article_number", "required": False, "max_len": 64, "type": "str"},
        {"key": "serial_number", "required": False, "max_len": 128, "type": "str"},
        {"key": "partner_name", "required": False, "max_len": 256, "type": "str"},  # szerződött partner
        {"key": "location_type", "required": False, "max_len": 64, "type": "str"},
        {"key": "counter", "required": False, "max_len": None, "type": "int"},
        {"key": "norm", "required": False, "max_len": None, "type": "float"},
        {"key": "tangible", "required": False, "max_len": None, "type": "bool"},
        {"key": "notes", "required": False, "max_len": None, "type": "str"},
    ],
}

EXPORT_ENTITIES = ("partners", "products", "assets", "settlements")


# ─── Fájl-parszolás ──────────────────────────────────────────────────────────


def _decode_file(data_url: str) -> tuple[bytes, str]:
    """data URL → (bytes, kind). kind: 'xlsx' | 'csv' (mime alapján)."""
    if not data_url.startswith("data:"):
        raise HTTPException(status_code=422, detail={"code": "impex.bad_file"})
    try:
        header, b64 = data_url.split(",", 1)
        raw = base64.b64decode(b64)
    except Exception:
        raise HTTPException(status_code=422, detail={"code": "impex.bad_file"})
    if len(raw) > _MAX_FILE_BYTES:
        raise HTTPException(status_code=422, detail={"code": "impex.file_too_large"})
    mime = header.split(";")[0].removeprefix("data:").lower()
    if "spreadsheet" in mime or "excel" in mime or raw[:2] == b"PK":
        return raw, "xlsx"
    return raw, "csv"


def _parse_rows(raw: bytes, kind: str, sheet: int = 0) -> tuple[list[list[str]], list[str]]:
    """A fájl sorai a kiválasztott munkalapról (fülről) + az összes fül neve.
    Minden cella stringgé alakítva ('' ha üres). CSV-nél egyetlen 'fül' van."""
    sheet_names: list[str] = []
    if kind == "xlsx":
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet_names = list(wb.sheetnames)
            if not (0 <= sheet < len(wb.worksheets)):
                wb.close()
                raise HTTPException(status_code=422, detail={"code": "impex.bad_sheet"})
            ws = wb.worksheets[sheet]

            def cell_str(c) -> str:
                if c is None:
                    return ""
                # Excel a számokat floatként adhatja — az egészeket ne
                # "11241539519.0" formában kapjuk (vonalkód, irányítószám!)
                if isinstance(c, float) and c.is_integer():
                    return str(int(c))
                return str(c).strip()

            rows = []
            for row in ws.iter_rows(max_row=_MAX_ROWS + 1, values_only=True):
                rows.append([cell_str(c) for c in row])
            wb.close()
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=422, detail={"code": "impex.parse_failed"})
    else:
        try:
            text = raw.decode("utf-8-sig", errors="replace")
            # pontosvessző vagy vessző — magyar Excel gyakran ;-t használ
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=";,\t")
            rows = [
                [c.strip() for c in row]
                for row in csv.reader(io.StringIO(text), dialect)
            ][: _MAX_ROWS + 1]
        except csv.Error:
            rows = [
                [c.strip() for c in row] for row in csv.reader(io.StringIO(text))
            ][: _MAX_ROWS + 1]
        except Exception:
            raise HTTPException(status_code=422, detail={"code": "impex.parse_failed"})
    # üres sorok kiszűrése
    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        raise HTTPException(status_code=422, detail={"code": "impex.empty_file"})
    return rows, sheet_names


def _column_letter(index: int) -> str:
    letters = ""
    index += 1
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


# ─── Végpontok ───────────────────────────────────────────────────────────────


@router.get("/entities")
async def list_entities(_: User = Depends(require_perm("import_export"))):
    return {
        "import": [
            {"entity": entity, "fields": fields}
            for entity, fields in ENTITY_FIELDS.items()
        ],
        "export": list(EXPORT_ENTITIES),
    }


class PreviewBody(BaseModel):
    file: str  # data URL
    sheet: int = Field(default=0, ge=0)  # munkalap (fül) indexe xlsx-nél


@router.post("/preview")
async def preview_file(
    body: PreviewBody,
    _: User = Depends(require_perm("import_export")),
):
    raw, kind = _decode_file(body.file)
    rows, sheet_names = _parse_rows(raw, kind, body.sheet)
    width = max(len(r) for r in rows)
    return {
        "columns": [
            {"index": i, "letter": _column_letter(i), "header": rows[0][i] if i < len(rows[0]) else ""}
            for i in range(width)
        ],
        "sample_rows": [r + [""] * (width - len(r)) for r in rows[1 : 1 + _PREVIEW_ROWS]],
        "total_rows": len(rows),
        "sheets": sheet_names,  # üres lista CSV-nél
        "sheet": body.sheet,
    }


class ImportBody(BaseModel):
    entity: str
    file: str  # data URL (ugyanaz, mint a preview-nál)
    mapping: dict[str, int]  # {mező_kulcs: oszlopindex}
    has_header: bool = True
    sheet: int = Field(default=0, ge=0)  # munkalap (fül) indexe xlsx-nél


def _coerce(value: str, field: dict) -> object:
    """Cellaérték → mezőérték a field típusa szerint. ValueError hibánál."""
    value = value.strip()
    if not value:
        return None
    if field["type"] == "int":
        return int(float(value.replace(",", ".").replace(" ", "")))
    if field["type"] == "float":
        return float(value.replace(",", ".").replace(" ", ""))
    if field["type"] == "bool":
        norm = value.lower()
        if norm in ("true", "igaz", "igen", "yes", "1", "x"):
            return True
        if norm in ("false", "hamis", "nem", "no", "0", ""):
            return False
        raise ValueError("bad bool")
    if field["type"] == "partner_type":
        norm = value.lower()
        aliases = {
            "customer": "customer", "vevő": "customer", "vevo": "customer",
            "supplier": "supplier", "szállító": "supplier", "szallito": "supplier",
            "both": "both", "mindkettő": "both", "mindketto": "both",
            "vevő és szállító": "both",
        }
        if norm not in aliases:
            raise ValueError("bad partner_type")
        return aliases[norm]
    if field["max_len"] and len(value) > field["max_len"]:
        value = value[: field["max_len"]]
    return value


@router.post("/import")
async def run_import(
    body: ImportBody,
    request: Request,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    actor: User = Depends(require_perm("import_export")),
):
    new_machine_combos: set[tuple[str, str]] = set()
    if body.entity not in ENTITY_FIELDS:
        raise HTTPException(status_code=422, detail={"code": "impex.bad_entity"})
    fields = {f["key"]: f for f in ENTITY_FIELDS[body.entity]}
    for key in body.mapping:
        if key not in fields:
            raise HTTPException(status_code=422, detail={"code": "impex.bad_mapping"})
    required = [f["key"] for f in ENTITY_FIELDS[body.entity] if f["required"]]
    for key in required:
        if key not in body.mapping:
            raise HTTPException(status_code=422, detail={"code": "impex.missing_required"})

    raw, kind = _decode_file(body.file)
    rows, _sheet_names = _parse_rows(raw, kind, body.sheet)
    data_rows = rows[1:] if body.has_header else rows
    if len(data_rows) > _MAX_ROWS:
        raise HTTPException(status_code=422, detail={"code": "impex.too_many_rows"})

    # Duplikátum-kihagyás kulcsa: gépeknél a vonalkód, egyébként a név.
    if body.entity == "assets":
        existing_keys = {
            b.strip().lower() for b in (await db.execute(select(Asset.barcode))).scalars()
        }
    else:
        model = Partner if body.entity == "partners" else Product
        existing_keys = {
            n.strip().lower() for n in (await db.execute(select(model.name))).scalars()
        }

    # ügyfél-kód sorszámozás (PT-NNNN) az importált partnereknek
    next_partner_num = 0
    if body.entity == "partners":
        from app.services.wfm.codes import _next_partner_number

        codes = list((await db.execute(select(Partner.partner_code))).scalars())
        next_partner_num = _next_partner_number(codes)

    # Gépeknél: partner-feloldás névre vagy PT-kódra ("szerződött partner")
    partners_by_key: dict[str, Partner] = {}
    if body.entity == "assets":
        for p in (await db.execute(select(Partner))).scalars():
            partners_by_key[p.name.strip().lower()] = p
            if p.partner_code:
                partners_by_key[p.partner_code.lower()] = p

    created = 0
    skipped = 0
    errors: list[dict] = []
    row_offset = 2 if body.has_header else 1  # Excel-sorszám a hibaüzenetben

    for i, row in enumerate(data_rows):
        values: dict[str, object] = {}
        row_error: str | None = None
        for key, col in body.mapping.items():
            cell = row[col] if col < len(row) else ""
            try:
                coerced = _coerce(cell, fields[key])
            except (ValueError, TypeError):
                row_error = f"{key}: '{cell}'"
                break
            if coerced is not None:
                values[key] = coerced
        if row_error:
            errors.append({"row": i + row_offset, "error": row_error})
            continue
        name = str(values.get("name") or "").strip()
        if not name:
            errors.append({"row": i + row_offset, "error": "name: üres"})
            continue
        key_field = "barcode" if body.entity == "assets" else "name"
        key = str(values.get(key_field) or "").strip()
        if not key:
            errors.append({"row": i + row_offset, "error": f"{key_field}: üres"})
            continue
        if key.lower() in existing_keys:
            skipped += 1
            continue
        existing_keys.add(key.lower())

        if body.entity == "assets":
            partner_ref = str(values.pop("partner_name", "") or "").strip()
            partner = partners_by_key.get(partner_ref.lower()) if partner_ref else None
            if partner_ref and partner is None:
                errors.append(
                    {"row": i + row_offset, "error": f"partner_name: '{partner_ref}' nem található"}
                )
                existing_keys.discard(key.lower())
                continue
            asset = Asset(**values, created_by=actor.id)
            new_machine_combos.add((asset.manufacturer or "", asset.name or ""))
            if partner is not None:
                asset.status = "deployed"
                asset.partner_id = partner.id
                asset.deployed_at = dt.now(UTC)
            db.add(asset)
            await db.flush()
            db.add(AssetMovement(asset_id=asset.id, action="created", actor_user_id=actor.id))
            if partner is not None:
                db.add(
                    AssetMovement(
                        asset_id=asset.id, action="deploy", partner_id=partner.id,
                        detail="import", actor_user_id=actor.id,
                    )
                )
            created += 1
            continue

        if body.entity == "partners":
            from app.api.inventory import compose_address

            composed = compose_address(
                values.get("address_zip"), values.get("address_city"),
                values.get("address_street"), values.get("address_number"),
            )
            if composed and not values.get("address"):
                values["address"] = composed
            composed_billing = compose_address(
                values.get("billing_zip"), values.get("billing_city"),
                values.get("billing_street"), values.get("billing_number"),
            )
            if composed_billing and not values.get("billing_address"):
                values["billing_address"] = composed_billing
            db.add(Partner(**values, partner_code=f"PT-{next_partner_num:04d}"))
            next_partner_num += 1
        else:
            db.add(Product(**values))
        created += 1

    await record_audit(
        db, actor=actor, action="impex.import", entity_type=body.entity,
        detail={"created": created, "skipped": skipped, "errors": len(errors)},
        request=request,
    )
    await db.commit()

    # Új géptípusoknál tudásbázis-bővítés a háttérben (importonként max. 10
    # kombináció, hogy a tömeges import ne indítson AI-hívás-lavinát).
    if new_machine_combos:
        from app.services.wfm.kb_auto import ensure_machine_kb

        for manufacturer, name in sorted(new_machine_combos)[:10]:
            background_tasks.add_task(ensure_machine_kb, manufacturer, name)

    return {"created": created, "skipped": skipped, "errors": errors[:100], "total": len(data_rows)}


# ─── Export ──────────────────────────────────────────────────────────────────


def _xlsx_response(headers: list[str], rows: list[list], filename: str) -> Response:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/{entity}")
async def run_export(
    entity: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    actor: User = Depends(require_perm("import_export")),
):
    if entity not in EXPORT_ENTITIES:
        raise HTTPException(status_code=422, detail={"code": "impex.bad_entity"})

    if entity == "partners":
        rows = (await db.execute(select(Partner).order_by(Partner.name))).scalars().all()
        headers = ["partner_code"] + [f["key"] for f in ENTITY_FIELDS["partners"]] + ["is_active"]
        data = [
            [getattr(p, h) for h in headers]
            for p in rows
        ]
    elif entity == "products":
        rows = (await db.execute(select(Product).order_by(Product.name))).scalars().all()
        headers = [f["key"] for f in ENTITY_FIELDS["products"]] + ["is_active"]
        data = [[getattr(p, h) for h in headers] for p in rows]
    elif entity == "assets":
        result = (
            await db.execute(
                select(Asset, Partner.name)
                .outerjoin(Partner, Partner.id == Asset.partner_id)
                .order_by(Asset.name)
            )
        ).all()
        headers = [
            "barcode", "manufacturer", "name", "article_number", "serial_number",
            "partner_name", "location_type", "counter", "norm", "tangible", "status",
        ]
        data = [
            [
                a.barcode, a.manufacturer, a.name, a.article_number, a.serial_number,
                pname, a.location_type, a.counter, a.norm,
                "True" if a.tangible else "False", a.status,
            ]
            for a, pname in result
        ]
    else:  # settlements
        result = (
            await db.execute(
                select(Settlement, Partner.name)
                .join(Partner, Partner.id == Settlement.partner_id)
                .order_by(Settlement.created_at.desc())
            )
        ).all()
        headers = [
            "date", "partner", "settled_by", "payment_method",
            "total_net", "total_gross", "invoiced", "billingo_document_id",
        ]
        data = [
            [
                s.created_at.strftime("%Y-%m-%d %H:%M"), pname, s.settled_by_name,
                s.payment_method, s.total_net, s.total_gross,
                "yes" if s.invoiced else "no", s.billingo_document_id or "",
            ]
            for s, pname in result
        ]

    await record_audit(
        db, actor=actor, action="impex.export", entity_type=entity,
        detail={"rows": len(data)}, request=request,
    )
    await db.commit()
    return _xlsx_response(headers, data, f"iwfm_{entity}.xlsx")
