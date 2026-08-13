"""Payroll export: aggregation correctness, CSV/XLSX content, audit, RBAC."""

import io
from datetime import UTC, date, datetime, time, timedelta

from openpyxl import load_workbook
from sqlalchemy import select

import app.db as app_db
from app.models import AuditEvent
from tests.conftest import (
    gen_bank_account,
    gen_tax_id,
    gen_taj,
)


async def seed_period(client, admin_headers, mgr_headers):
    """One employee with: 5×8h published shifts, one 8h time entry, 2 leave days."""
    created = await client.post(
        "/api/employees",
        json={
            "email": "ber.dolgozo@example.com",
            "last_name": "Bér",
            "first_name": "Béla",
            "hire_date": "2026-01-01",
            "feor_code": "9223",
            "job_title": "raktáros",
            "tax_id": gen_tax_id(3),
            "taj": gen_taj(3),
            "bank_account": gen_bank_account(3),
        },
        headers=admin_headers,
    )
    emp_id = created.json()["id"]

    day = date.today() + timedelta(days=21)
    monday = day - timedelta(days=day.weekday())
    for i in range(5):
        await client.post(
            "/api/shifts",
            json={
                "employee_id": emp_id,
                "work_date": (monday + timedelta(days=i)).isoformat(),
                "start_time": "08:00:00",
                "end_time": "16:30:00",
                "break_minutes": 30,
            },
            headers=mgr_headers,
        )
    await client.post(
        "/api/shifts/publish", json={"week_start": monday.isoformat()}, headers=mgr_headers
    )

    clock_in = datetime.combine(monday, time(8, 0), tzinfo=UTC)
    await client.post(
        "/api/time-entries",
        json={
            "employee_id": emp_id,
            "clock_in": clock_in.isoformat(),
            "clock_out": (clock_in + timedelta(hours=8, minutes=30)).isoformat(),
            "break_minutes": 30,
        },
        headers=mgr_headers,
    )

    leave_start = monday + timedelta(days=7)
    rid = (
        await client.post(
            "/api/time-off",
            json={
                "employee_id": emp_id,
                "type": "annual",
                "start_date": leave_start.isoformat(),
                "end_date": (leave_start + timedelta(days=1)).isoformat(),
            },
            headers=mgr_headers,
        )
    ).json()["id"]
    await client.patch(
        f"/api/time-off/{rid}/decide", json={"status": "approved"}, headers=mgr_headers
    )
    return monday


async def test_csv_export_content(client, admin, manager):
    _, admin_headers = admin
    _, mgr_headers = manager
    monday = await seed_period(client, admin_headers, mgr_headers)

    res = await client.get(
        "/api/payroll/export",
        params={
            "period_start": monday.isoformat(),
            "period_end": (monday + timedelta(days=13)).isoformat(),
            "format": "csv",
        },
        headers=mgr_headers,
    )
    assert res.status_code == 200
    assert "text/csv" in res.headers["content-type"]
    text = res.content.decode("utf-8-sig")
    assert "Bér Béla" in text
    assert gen_tax_id(3) in text  # dekriptált azonosító az exportban
    # 5×8h beosztott, 8h ledolgozott, 2 nap szabadság
    row = next(line for line in text.splitlines() if line.startswith("Bér Béla"))
    cells = row.split(";")
    assert cells[7] == "40.0"  # beosztott óra
    assert cells[8] == "8.0"  # ledolgozott óra
    assert cells[9] == "2"  # szabadság nap


async def test_xlsx_export_content(client, admin, manager):
    _, admin_headers = admin
    _, mgr_headers = manager
    monday = await seed_period(client, admin_headers, mgr_headers)

    res = await client.get(
        "/api/payroll/export",
        params={
            "period_start": monday.isoformat(),
            "period_end": (monday + timedelta(days=13)).isoformat(),
            "format": "xlsx",
        },
        headers=mgr_headers,
    )
    assert res.status_code == 200
    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws["A2"].value == "Dolgozó neve"
    assert ws["A3"].value == "Bér Béla"
    assert ws["H3"].value == 40.0


async def test_export_audited(client, admin, manager):
    _, admin_headers = admin
    _, mgr_headers = manager
    monday = await seed_period(client, admin_headers, mgr_headers)
    await client.get(
        "/api/payroll/export",
        params={
            "period_start": monday.isoformat(),
            "period_end": (monday + timedelta(days=6)).isoformat(),
        },
        headers=mgr_headers,
    )
    factory = app_db.get_session_factory()
    async with factory() as session:
        actions = (await session.execute(select(AuditEvent.action))).scalars().all()
    assert "payroll.export" in actions


async def test_export_validation_and_rbac(client, manager, employee_user):
    _, mgr_headers = manager
    _, emp_headers, _ = employee_user
    today = date.today()

    # fordított időszak
    assert (
        await client.get(
            "/api/payroll/export",
            params={
                "period_start": today.isoformat(),
                "period_end": (today - timedelta(days=5)).isoformat(),
            },
            headers=mgr_headers,
        )
    ).status_code == 422
    # túl hosszú időszak
    assert (
        await client.get(
            "/api/payroll/export",
            params={
                "period_start": today.isoformat(),
                "period_end": (today + timedelta(days=200)).isoformat(),
            },
            headers=mgr_headers,
        )
    ).status_code == 422
    # dolgozó nem exportálhat
    assert (
        await client.get(
            "/api/payroll/export",
            params={
                "period_start": today.isoformat(),
                "period_end": (today + timedelta(days=6)).isoformat(),
            },
            headers=emp_headers,
        )
    ).status_code == 403
