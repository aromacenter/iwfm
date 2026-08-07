"""Import/export: preview, oszlop-hozzárendeléses import, dedup, hibasorok, export."""

import base64
import io


def xlsx_data_url(rows: list[list]) -> str:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    b64 = base64.b64encode(buf.getvalue()).decode()
    return f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"


PARTNER_SHEET = [
    ["Cégnév", "Kapcsolattartó", "Email", "Típus", "Fiz. határidő"],
    ["Kávézó Bt.", "Kis Pál", "pal@kavezo.hu", "Vevő", "8"],
    ["Hotel Zrt.", "Nagy Éva", "eva@hotel.hu", "vevő", "30"],
    ["", "", "", "", ""],  # üres sor — kiszűrve
    ["Rossz Sor Kft.", "X", "x@x.hu", "nemtudom", "5"],  # hibás típus
]


async def test_preview(client, manager):
    _, mgr = manager
    res = await client.post(
        "/api/import-export/preview", json={"file": xlsx_data_url(PARTNER_SHEET)}, headers=mgr
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["columns"][0]["letter"] == "A"
    assert body["columns"][0]["header"] == "Cégnév"
    assert body["total_rows"] == 4  # üres sor kiszűrve
    assert body["sample_rows"][0][0] == "Kávézó Bt."


async def test_import_partners_with_mapping(client, manager):
    _, mgr = manager
    res = await client.post(
        "/api/import-export/import",
        json={
            "entity": "partners",
            "file": xlsx_data_url(PARTNER_SHEET),
            "mapping": {
                "name": 0,
                "contact_name": 1,
                "contact_email": 2,
                "partner_type": 3,
                "payment_terms_days": 4,
            },
        },
        headers=mgr,
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["created"] == 2
    assert len(body["errors"]) == 1  # a "nemtudom" típusú sor
    assert body["errors"][0]["error"].startswith("partner_type")

    partners = (await client.get("/api/partners", headers=mgr)).json()
    names = {p["name"]: p for p in partners}
    assert "Kávézó Bt." in names and "Hotel Zrt." in names
    assert names["Kávézó Bt."]["partner_type"] == "customer"
    assert names["Hotel Zrt."]["payment_terms_days"] == 30

    # ismételt import: a meglévők kihagyva; a korábban hibás sor (most csak
    # név-mapping → érvényes) létrejön
    again = await client.post(
        "/api/import-export/import",
        json={
            "entity": "partners",
            "file": xlsx_data_url(PARTNER_SHEET),
            "mapping": {"name": 0},
        },
        headers=mgr,
    )
    assert again.json()["created"] == 1  # Rossz Sor Kft.
    assert again.json()["skipped"] == 2


async def test_import_products(client, manager):
    _, mgr = manager
    sheet = [
        ["Termék", "g/adag", "Ár"],
        ["Házi kávé", "7", "55,5"],  # magyar tizedesvessző
        ["Prémium kávé", "8", "80"],
    ]
    res = await client.post(
        "/api/import-export/import",
        json={
            "entity": "products",
            "file": xlsx_data_url(sheet),
            "mapping": {"name": 0, "grams_per_portion": 1, "price_per_portion": 2},
        },
        headers=mgr,
    )
    assert res.status_code == 200, res.text
    assert res.json()["created"] == 2
    products = (await client.get("/api/products", headers=mgr)).json()
    by_name = {p["name"]: p for p in products}
    assert by_name["Házi kávé"]["price_per_portion"] == 55.5
    assert by_name["Prémium kávé"]["grams_per_portion"] == 8


async def test_import_requires_name_mapping(client, manager):
    _, mgr = manager
    res = await client.post(
        "/api/import-export/import",
        json={
            "entity": "partners",
            "file": xlsx_data_url(PARTNER_SHEET),
            "mapping": {"contact_name": 1},
        },
        headers=mgr,
    )
    assert res.status_code == 422
    assert res.json()["detail"]["code"] == "impex.missing_required"


async def test_export_partners_xlsx(client, manager):
    _, mgr = manager
    await client.post(
        "/api/partners", json={"name": "Export Kft."}, headers=mgr
    )
    res = await client.get("/api/import-export/export/partners", headers=mgr)
    assert res.status_code == 200
    assert res.content[:2] == b"PK"  # xlsx = zip
    assert "iwfm_partners.xlsx" in res.headers["content-disposition"]

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(res.content))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0][0] == "name"
    assert any(r[0] == "Export Kft." for r in rows[1:])


async def test_import_export_rbac(client, employee_user):
    _, emp_headers, _ = employee_user
    assert (await client.get("/api/import-export/entities", headers=emp_headers)).status_code == 403
    assert (
        await client.get("/api/import-export/export/partners", headers=emp_headers)
    ).status_code == 403
